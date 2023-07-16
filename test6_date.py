from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the workbook
workbook = load_workbook('destination_file.xlsx')
source_sheet1 = workbook['dam-2']
source_sheet2 = workbook['rtm-2']
source_sheet3= workbook['dsm-1']

start_date = datetime(2023, 2, 10).date()
end_date = datetime(2023, 5, 15).date()
global dates 
dates = []

for single_date in range((end_date - start_date).days + 1):
    current_date = start_date + timedelta(days=single_date)
    formatted_date = current_date.strftime("%d.%m.%y")
    dates.append(formatted_date)
    # print(formatted_date)
print(dates)

def data_col(source_sheet,j):
    for i in dates:
        heading = i
        if j == "A":
            new_sheet = workbook.create_sheet(title=i)
            column_letter = None
            for cell in source_sheet[1]:
                if cell.value == heading:
                    column_letter = get_column_letter(cell.column)
            if column_letter:
                source_column = source_sheet[column_letter]
                for index, cell in enumerate(source_column, start=1):
                    # new_sheet.cell(row=index, column=j, value=cell.value)
                    new_sheet[j + str(index)].value = cell.value
        else:
            new_sheet = workbook[i]
            column_letter = None
            for cell in source_sheet[1]:
                if cell.value == heading:
                    column_letter = get_column_letter(cell.column)
            if column_letter:
                source_column = source_sheet[column_letter]
                for index, cell in enumerate(source_column, start=1):
                    # new_sheet.cell(row=index, column=j, value=cell.value)
                    new_sheet[j + str(index)].value = cell.value
          
data_col(source_sheet1,"A")
print("A done")
data_col(source_sheet2,"B")
print("B done")
data_col(source_sheet3,"C")
print("C done")

           
workbook.save('dest2.xlsx')