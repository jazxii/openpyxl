import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('destination_file.xlsx')

# Select the desired sheet by name or index
sheet = workbook['dsm-2']  # Replace 'Sheet1' with the name of your sheet

# Get the maximum column index
max_col = sheet.max_column

# Iterate over the columns in reverse order
for col in range(max_col, 0, -1):
    column_values = sheet[sheet.cell(row=1, column=col).column_letter]

    # Check if all cells in the column are empty
    if all(cell.value is None for cell in column_values):
        sheet.delete_cols(col)

# Save the modified workbook
workbook.save('destination_file.xlsx')