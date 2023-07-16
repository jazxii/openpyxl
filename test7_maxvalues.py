from openpyxl import load_workbook
from datetime import datetime, timedelta


# Load the workbook
workbook = load_workbook('dest2.xlsx')


start_date = datetime(2023, 2, 10).date()
end_date = datetime(2023, 5, 15).date()
global dates 
dates = []

for single_date in range((end_date - start_date).days + 1):
    current_date = start_date + timedelta(days=single_date)
    formatted_date = current_date.strftime("%d.%m.%y")
    dates.append(formatted_date)
    # print(formatted_date)
print(dates)
x = 400
# Iterate over each sheet
for i in dates:
    sheet = workbook[i]
    
    # Iterate over each row
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Check if any value in the row is negative
        if any(value < 0 for value in row):
            # If any value is negative, set column D to zero
            sheet.cell(row=row_index, column=4).value = 0
        else:
            # If all values are non-negative, set column D to the maximum value
            max_value = max(row[0], row[1], row[2])
            sheet.cell(row=row_index, column=4).value = max_value
            # sheet.cell(row=row_index, column=5).value = max_value*x
        sheet.cell(row=1, column=4).value = 'Max Value'




# Save the workbook
workbook.save('dest3_compared.xlsx')  # Replace 'your_updated_workbook.xlsx' with the desired filename
