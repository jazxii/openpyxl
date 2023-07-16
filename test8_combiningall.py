from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from datetime import datetime, timedelta


# Load the workbook
workbook = load_workbook('dest3_compared.xlsx')


start_date = datetime(2023, 2, 10).date()
end_date = datetime(2023, 5, 15).date()
global dates 
dates = []

for single_date in range((end_date - start_date).days + 1):
    current_date = start_date + timedelta(days=single_date)
    formatted_date = current_date.strftime("%d.%m.%y")
    dates.append(formatted_date)
    # print(formatted_date)
print(dates)

new_sheet = workbook.create_sheet(title='Combined max rates')
no = len(dates)

column_index = 1

for i in dates:
    source_sheet = workbook[i]
    heading = 'Max Value'
    column_letter = None
    for cell in source_sheet[1]:
        if cell.value == heading:
            column_letter = get_column_letter(cell.column)
        if column_letter:
            source_column = source_sheet[column_letter]
            target_col = get_column_letter(column_index)
            for index, cell in enumerate(source_column, start=1):
                # new_sheet.cell(row=index, column=j, value=cell.value)
                new_sheet[target_col + str(index)].value = cell.value
    column_index += 1
    


workbook.save('combined.xlsx')
    

