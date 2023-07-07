import openpyxl

def copy_rows_to_columns(source_sheet, source_rows, destination_sheet, destination_start_column):
    # Open the source workbook
    source_workbook = openpyxl.load_workbook('Book2.xlsx')

    # Get the source sheet
    source_sheet = source_workbook[source_sheet]

    # Get the values from the source rows
    source_values = []
    for row in source_rows:
        row_values = [cell.value for cell in source_sheet[row]]
        source_values.append(row_values)

    # Open the destination workbook
    destination_workbook = openpyxl.load_workbook('Book2.xlsx')

    # Get the destination sheet
    destination_sheet = destination_workbook[destination_sheet]

    # Paste the values as columns in the destination sheet
    for row_index, row_values in enumerate(source_values):
        for col_index, value in enumerate(row_values):
            destination_sheet.cell(row=row_index + 1, column=destination_start_column + col_index).value = value

    # Save the destination workbook
    destination_workbook.save('destination_file.xlsx')

# Example usage
copy_rows_to_columns('dam', [2, 5, 8], 'Finale', 1)