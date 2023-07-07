import openpyxl

def copy_rows_to_columns(source_sheet, source_rows, destination_sheet, destination_start_column):
    # Open the source workbook
    source_workbook = openpyxl.load_workbook('Book2.xlsx')

    # Get the source sheet
    source_sheet = source_workbook[source_sheet]

    # Open the destination workbook
    destination_workbook = openpyxl.load_workbook('destination_file.xlsx')

    # Get the destination sheet
    destination_sheet = destination_workbook[destination_sheet]

    # Paste the transposed rows as columns in the destination sheet
    for row_index, row_number in enumerate(source_rows):
        row_values = [cell.value for cell in source_sheet[row_number]]
        for col_index, value in enumerate(row_values):
            destination_sheet.cell(row=col_index + 1, column=destination_start_column + row_index).value = value

    # Save the destination workbook
    destination_workbook.save('destination_file.xlsx')


array = list(range(122, 404, 3))
array1 = list(range(123,405,3))

# Example usage
copy_rows_to_columns('dam', array, 'dam-1', 1)
copy_rows_to_columns('dam', array1, 'dam-2', 1)
copy_rows_to_columns('rtm', array, 'rtm-1', 1)
copy_rows_to_columns('rtm', array1, 'rtm-2', 1)