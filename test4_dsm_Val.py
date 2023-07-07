import openpyxl

def copy_rows_to_columns(source_sheet, source_column, destination_sheet):
    # Open the source workbook
    source_workbook = openpyxl.load_workbook('Book2.xlsx')

    # Get the source sheet
    source_sheet = source_workbook[source_sheet]

    # Open the destination workbook
    destination_workbook = openpyxl.load_workbook('destination_file.xlsx')

    # Get the destination sheet
    destination_sheet = destination_workbook[destination_sheet]

    # Copy and paste rows in groups of 96
    row_count = source_sheet.max_row
    for start_row in range(1, row_count + 1, 96):
        end_row = min(start_row + 95, row_count)  # Limit end row to avoid going beyond the last row

        # Copy rows from source sheet
        source_values = [cell.value for row in source_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=source_column, max_col=source_column) for cell in row]

        # Calculate the destination column index based on the group number
        group_number = (start_row - 1) // 96
        destination_column = destination_sheet.max_column + group_number + 1

        # Paste rows as a column in the destination sheet
        for row_index, value in enumerate(source_values):
            destination_sheet.cell(row=row_index + 1, column=destination_column).value = value

    # Save the destination workbook
    destination_workbook.save('destination_file.xlsx')

# Example usage
copy_rows_to_columns('dsm', 2, 'dsm-1')
copy_rows_to_columns('dsm', 3, 'dsm-2')
